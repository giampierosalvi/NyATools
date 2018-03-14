""" Compares two exel files in a specific way described below. This is to take care of the fact that
    the admission office adds new applications to NyA and sometimes changes the personal number of
    existing applicants.

    python smartCompare.py oldNyA.xlsx newNyA.xlsx diff.xlsx

    Compares lists of personal numbers and returs:
    1) the ones in oldNyA.xlsx but not in newNyA.xlsx and
    2) the ones in newNyA.xlsx but not in oldNyA.xlsx.
    in both cases, it checks first and family names and tries to find out if there was a change of
    personal number.
    Finally it saves the rows to be added to your working spreadsheet to diff.xlsx

    NOTE: NyA saves in xls format: you need to open in libreoffice and save as xlsx

    (C) 2016-2018 Giampiero Salvi <giampi@kth.se>
"""
import sys
from openpyxl import load_workbook, Workbook
import re
#from sets import Set

def worksheet2array(ws, header=True):
    """ converts worksheet object into array. If header=True returns first row in a separate array """
    head = None
    wsarray = []
    for row in ws.rows:
        rowarray = []
        for cell in row:
            rowarray.append(cell.value)
        wsarray.append(rowarray)
    if header:
        head = wsarray[0]
        wsarray = wsarray[1:]
    return wsarray, head

def parseAdditionalInfo(info):
    """ parse the additional info, not reliying entirely on the comma separated values """
    infoArray = [s.strip() for s in info.split(',')]
    idx=0
    bachelorList = []
    while(re.match(r'^[0-9.]*$', infoArray[idx])==None):
        bachelorList.append(infoArray[idx])
        idx = idx+1
    bachelor = ', '.join(bachelorList)
    nCredits = infoArray[idx]
    date = infoArray[idx+1]
    country = infoArray[idx+2]
    rest = ', '.join(infoArray[(idx+3):])
    return [bachelor, nCredits, date, country, rest]

oldFileName = sys.argv[1]
newFileName = sys.argv[2]
outFileName = sys.argv[3]

# load workbooks
oldWorkBook = load_workbook(oldFileName)
newWorkBook = load_workbook(newFileName)

# select active worksheet
oldWorkSheet = oldWorkBook.active
newWorkSheet = newWorkBook.active
# examples of use
# newWorkSheet['A1'].value
# newWorkSheet.columns

# convert into array
oldData, oldHead = worksheet2array(oldWorkSheet)
newData, newHead = worksheet2array(newWorkSheet)

# get list of personal numbers
oldPN = [a[0] for a in oldData]
newPN = [a[0] for a in newData]

# convert to sets
oldPNSet = set(oldPN)
newPNSet = set(newPN)

# find differences
toremove = oldPNSet.difference(newPNSet)
toadd = newPNSet.difference(oldPNSet)

# extract names concatenating first and surname
oldNames = [oldData[i][1]+'_'+oldData[i][2] for i in range(len(oldData))]
newNames = [newData[i][1]+'_'+newData[i][2] for i in range(len(newData))]
print("\ncheck for duplicate names in google spreadsheet")
print(set([x for x in oldNames if oldNames.count(x) > 1]))

print("\ncheck for duplicate names in NyA spreadsheet")
print(set([x for x in newNames if newNames.count(x) > 1]))

# check if different personal numbers correspond to the same name
toremovenames = [oldNames[oldPN.index(pn)] for pn in toremove]
toaddnames = [newNames[newPN.index(pn)] for pn in toadd]
tochangepn = set(toremovenames).intersection(set(toaddnames))
print("\nTo change personal number:")
for name in tochangepn:
    print(name)
    thisOldPN = oldPN[oldNames.index(name)]
    thisNewPN = newPN[newNames.index(name)]
    print(thisOldPN + ' -> ' + thisNewPN)
    toremove.remove(thisOldPN)
    toadd.remove(thisNewPN)

print("\nTo add according to personal number (with double check of names)")
toaddidx = []
for pn in toadd:
    i = newPN.index(pn)
    name = newData[i][1] + '_' + newData[i][2]
    print(newData[i][0], newData[i][1], newData[i][2])
    toaddidx.append(i)

print("\nTo remove according to personal number (with double check of names)")
for pn in toremove:
    i = oldPN.index(pn)
    name = oldData[i][1] + '_' + oldData[i][2]
    print(oldData[i][0], oldData[i][1], oldData[i][2])

# create output spreadsheet
addinfoidx = newHead.index('ADDITIONAL INFO')
outwb = Workbook()
outws = outwb.active
for n in range(len(toaddidx)):
    i = toaddidx[n]
    rowarray = newData[i]
    if rowarray[addinfoidx]!=None:
        infoarray = parseAdditionalInfo(rowarray[addinfoidx])
        if len(infoarray)>1:
            rowarray[addinfoidx:(addinfoidx+len(infoarray)-1)] = infoarray
    for j in range(len(rowarray)):
        cellname = chr(j + ord('A'))+str(n+1)
        outws[cellname] = rowarray[j]

outwb.save(outFileName)
